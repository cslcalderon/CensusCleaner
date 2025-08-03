from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

def apply_cond_format (sheet, standard_columns, max_row = 1000):
    def get_col_letter_by_name(name):
        try: 
            indx = standard_columns.index(name) + 1
            return get_column_letter(indx)
        except ValueError:
            return None
        
    rel_col = get_col_letter_by_name("Relationship")
    tier_col = get_col_letter_by_name("Tier")

    #relationship rules and formatting 
    if rel_col: 
        cell_range = f"{rel_col}2:{rel_col}{max_row}" #ex: B2:B1000
        relationship_colors = {
            "Child": ("FFC7CE", "9C0006"), 
            "Employee": ("C6DCFD", "003366"), 
            "Spouse": ("FFEB9C", "9C6500")
        }

        for relation, (fill_color, font_color) in relationship_colors.items():
            sheet.conditional_formatting.add(
                cell_range, 
                FormulaRule(
                    formula=[f'{rel_col}2="{relation}"'],
                    fill=PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid"), 
                    font=Font(color = font_color)
                )
            )
    if tier_col: 
        cell_range = f"{tier_col}2:{tier_col}{max_row}" #ex: B2:B1000
        tier_colors = {
            "EC": ("FFC7CE", "9C0006"), 
            "EE": ("C6DCFD", "003366"), 
            "ES": ("FFEB9C", "9C6500"), 
            "EF": ("C6EFCE", "006100")
        }

        for relation, (fill_color, font_color) in tier_colors.items():
            sheet.conditional_formatting.add(
                cell_range, 
                FormulaRule(
                    formula=[f'{tier_col}2="{relation}"'],
                    fill=PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid"), 
                    font=Font(color = font_color)
                )
            )

       
       